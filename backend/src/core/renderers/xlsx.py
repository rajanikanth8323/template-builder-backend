# # src/core/renderers/xlsx.py
# # XLSX renderer — converts template layout to Excel file

# import io
# from typing import Any, Dict, List

# try:
#     import openpyxl
#     from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
#     HAS_OPENPYXL = True
# except ImportError:
#     HAS_OPENPYXL = False


# def _resolve(text: str, values: Dict[str, str]) -> str:
#     if not text:
#         return ""
#     for key, val in values.items():
#         text = text.replace(f"{{{{{key}}}}}", val)
#     return text


# class XlsxRenderer:
#     def render(self, layout_json: Dict[str, Any], context: Dict[str, Any]) -> bytes:
#         if not HAS_OPENPYXL:
#             raise RuntimeError("openpyxl is not installed. Run: pip install openpyxl")

#         values   = context.get("values", {})
#         datasets = context.get("datasets", {})
#         blocks   = layout_json.get("blocks", [])

#         wb = openpyxl.Workbook()
#         ws = wb.active
#         ws.title = "Document"

#         # Style helpers
#         header_font    = Font(bold=True, size=12, color="FFFFFF")
#         header_fill    = PatternFill("solid", fgColor="4F46E5")
#         header_align   = Alignment(horizontal="center", vertical="center")
#         row_fill_even  = PatternFill("solid", fgColor="F8FAFC")
#         row_fill_odd   = PatternFill("solid", fgColor="FFFFFF")
#         border_side    = Side(style="thin", color="E2E8F0")
#         cell_border    = Border(
#             left=border_side, right=border_side,
#             top=border_side, bottom=border_side
#         )

#         current_row = 1

#         for block in blocks:
#             btype = block.get("type", "")

#             # ── TEXT block ────────────────────────────────────────
#             if btype == "text":
#                 content = _resolve(block.get("content", ""), values)
#                 for line in content.split("\n"):
#                     if line.strip():
#                         cell = ws.cell(row=current_row, column=1, value=line.strip())
#                         cell.font      = Font(size=11)
#                         cell.alignment = Alignment(wrap_text=True)
#                         ws.merge_cells(
#                             start_row=current_row, start_column=1,
#                             end_row=current_row, end_column=5
#                         )
#                         ws.row_dimensions[current_row].height = 20
#                         current_row += 1
#                 current_row += 1  # blank row after text

#             # ── SECTION block ─────────────────────────────────────
#             elif btype == "section":
#                 label = _resolve(block.get("content", "Section"), values)
#                 cell = ws.cell(row=current_row, column=1, value=label.upper())
#                 cell.font      = Font(bold=True, size=11, color="4C1D95")
#                 cell.fill      = PatternFill("solid", fgColor="EDE9FE")
#                 cell.alignment = Alignment(horizontal="left", vertical="center")
#                 ws.merge_cells(
#                     start_row=current_row, start_column=1,
#                     end_row=current_row, end_column=5
#                 )
#                 ws.row_dimensions[current_row].height = 22
#                 current_row += 1

#             # ── TABLE block ───────────────────────────────────────
#             elif btype == "table":
#                 columns    = block.get("columns", [])
#                 block_id   = block.get("block_id", "")
#                 dataset    = datasets.get(block_id, [])

#                 if not columns:
#                     continue

#                 # Header row
#                 for ci, col in enumerate(columns, start=1):
#                     cell = ws.cell(row=current_row, column=ci, value=col.get("header", f"Col {ci}"))
#                     cell.font      = header_font
#                     cell.fill      = header_fill
#                     cell.alignment = header_align
#                     cell.border    = cell_border
#                     ws.column_dimensions[
#                         openpyxl.utils.get_column_letter(ci)
#                     ].width = 20
#                 ws.row_dimensions[current_row].height = 22
#                 current_row += 1

#                 # Data rows from dataset
#                 if dataset:
#                     for ri, data_row in enumerate(dataset):
#                         fill = row_fill_even if ri % 2 == 0 else row_fill_odd
#                         for ci, col in enumerate(columns, start=1):
#                             val = data_row[ci - 1] if ci - 1 < len(data_row) else ""
#                             cell = ws.cell(row=current_row, column=ci, value=val)
#                             cell.fill      = fill
#                             cell.border    = cell_border
#                             cell.alignment = Alignment(vertical="center")
#                         ws.row_dimensions[current_row].height = 18
#                         current_row += 1
#                 else:
#                     # Use binding row — resolve {{tokens}} with actual values
#                     for ci, col in enumerate(columns, start=1):
#                         binding = _resolve(col.get("binding", ""), values)
#                         cell = ws.cell(row=current_row, column=ci, value=binding)
#                         cell.fill   = row_fill_even
#                         cell.border = cell_border
#                     current_row += 1
#                     # Also add any manual rows
#                     for ri, data_row in enumerate(block.get("rows", [])):
#                         fill = row_fill_even if ri % 2 == 0 else row_fill_odd
#                         for ci, col in enumerate(columns, start=1):
#                             val = data_row[ci-1] if ci-1 < len(data_row) and data_row[ci-1] else ""
#                             resolved = _resolve(val, values) if val else _resolve(col.get("binding",""), values)
#                             cell = ws.cell(row=current_row, column=ci, value=resolved)
#                             cell.fill   = fill
#                             cell.border = cell_border
#                         current_row += 1

#                 current_row += 1  # blank row after table

#         # Auto-fit column widths
#         for col in ws.columns:
#             max_len = 0
#             col_letter = openpyxl.utils.get_column_letter(col[0].column)
#             for cell in col:
#                 try:
#                     if cell.value:
#                         max_len = max(max_len, len(str(cell.value)))
#                 except Exception:
#                     pass
#             ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 50)

#         # Save to bytes
#         buf = io.BytesIO()
#         wb.save(buf)
#         return buf.getvalue()

# src/core/renderers/xlsx.py
# XLSX renderer — converts template layout to Excel file

import io
import base64
import urllib.request
import tempfile
import os
from typing import Any, Dict, List

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XLImage
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def _resolve(text: str, values: Dict[str, str]) -> str:
    if not text:
        return ""
    for key, val in values.items():
        text = text.replace(f"{{{{{key}}}}}", val)
    return text


class XlsxRenderer:
    def render(self, layout_json: Dict[str, Any], context: Dict[str, Any]) -> bytes:
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl is not installed. Run: pip install openpyxl")

        values   = context.get("values", {})
        datasets = context.get("datasets", {})
        blocks   = layout_json.get("blocks", [])

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Document"

        # Style helpers
        header_font    = Font(bold=True, size=12, color="FFFFFF")
        header_fill    = PatternFill("solid", fgColor="4F46E5")
        header_align   = Alignment(horizontal="center", vertical="center")
        row_fill_even  = PatternFill("solid", fgColor="F8FAFC")
        row_fill_odd   = PatternFill("solid", fgColor="FFFFFF")
        border_side    = Side(style="thin", color="E2E8F0")
        cell_border    = Border(
            left=border_side, right=border_side,
            top=border_side, bottom=border_side
        )

        current_row = 1

        for block in blocks:
            btype = block.get("type", "")

            # ── TEXT block ────────────────────────────────────────
            if btype == "text":
                content = _resolve(block.get("content", ""), values)
                for line in content.split("\n"):
                    if line.strip():
                        cell = ws.cell(row=current_row, column=1, value=line.strip())
                        cell.font      = Font(size=11)
                        cell.alignment = Alignment(wrap_text=True)
                        ws.merge_cells(
                            start_row=current_row, start_column=1,
                            end_row=current_row, end_column=5
                        )
                        ws.row_dimensions[current_row].height = 20
                        current_row += 1
                current_row += 1  # blank row after text

            # ── SECTION block ─────────────────────────────────────
            elif btype == "section":
                label = _resolve(block.get("content", "Section"), values)
                cell = ws.cell(row=current_row, column=1, value=label.upper())
                cell.font      = Font(bold=True, size=11, color="4C1D95")
                cell.fill      = PatternFill("solid", fgColor="EDE9FE")
                cell.alignment = Alignment(horizontal="left", vertical="center")
                ws.merge_cells(
                    start_row=current_row, start_column=1,
                    end_row=current_row, end_column=5
                )
                ws.row_dimensions[current_row].height = 22
                current_row += 1

            # ── TABLE block ───────────────────────────────────────
            elif btype == "table":
                columns    = block.get("columns", [])
                block_id   = block.get("block_id", "")
                dataset    = datasets.get(block_id, [])

                if not columns:
                    continue

                # Header row
                for ci, col in enumerate(columns, start=1):
                    cell = ws.cell(row=current_row, column=ci, value=col.get("header", f"Col {ci}"))
                    cell.font      = header_font
                    cell.fill      = header_fill
                    cell.alignment = header_align
                    cell.border    = cell_border
                    ws.column_dimensions[
                        openpyxl.utils.get_column_letter(ci)
                    ].width = 20
                ws.row_dimensions[current_row].height = 22
                current_row += 1

                # Data rows from dataset
                if dataset:
                    for ri, data_row in enumerate(dataset):
                        fill = row_fill_even if ri % 2 == 0 else row_fill_odd
                        for ci, col in enumerate(columns, start=1):
                            val = data_row[ci - 1] if ci - 1 < len(data_row) else ""
                            cell = ws.cell(row=current_row, column=ci, value=val)
                            cell.fill      = fill
                            cell.border    = cell_border
                            cell.alignment = Alignment(vertical="center")
                        ws.row_dimensions[current_row].height = 18
                        current_row += 1
                else:
                    # Use binding row — resolve {{tokens}} with actual values
                    for ci, col in enumerate(columns, start=1):
                        binding = _resolve(col.get("binding", ""), values)
                        cell = ws.cell(row=current_row, column=ci, value=binding)
                        cell.fill   = row_fill_even
                        cell.border = cell_border
                    current_row += 1
                    # Also add any manual rows
                    for ri, data_row in enumerate(block.get("rows", [])):
                        fill = row_fill_even if ri % 2 == 0 else row_fill_odd
                        for ci, col in enumerate(columns, start=1):
                            val = data_row[ci-1] if ci-1 < len(data_row) and data_row[ci-1] else ""
                            resolved = _resolve(val, values) if val else _resolve(col.get("binding",""), values)
                            cell = ws.cell(row=current_row, column=ci, value=resolved)
                            cell.fill   = fill
                            cell.border = cell_border
                        current_row += 1

                current_row += 1  # blank row after table

            # ── IMAGE block ───────────────────────────────────────
            elif btype == "image":
                src = _resolve(block.get("src", ""), values)
                if src:
                    try:
                        img_bytes = self._load_image_bytes(src)
                        if img_bytes:
                            # Detect extension for temp file
                            suffix = ".png"
                            if "jpeg" in src[:50] or "jpg" in src[:50]:
                                suffix = ".jpg"
                            elif "gif" in src[:50]:
                                suffix = ".gif"
                            elif "webp" in src[:50]:
                                suffix = ".webp"

                            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                                tmp.write(img_bytes)
                                tmp_path = tmp.name

                            try:
                                xl_img = XLImage(tmp_path)
                                # Cap image size to reasonable dimensions
                                xl_img.width  = min(xl_img.width,  500)
                                xl_img.height = min(xl_img.height, 300)
                                cell_ref = f"A{current_row}"
                                ws.add_image(xl_img, cell_ref)
                                # Reserve rows for the image (~20px per row)
                                rows_needed = max(1, xl_img.height // 20)
                                current_row += rows_needed + 1
                            finally:
                                os.unlink(tmp_path)
                        else:
                            cell = ws.cell(row=current_row, column=1, value=f"[Image not found: {src}]")
                            cell.font = Font(color="94A3B8", italic=True, size=10)
                            current_row += 2
                    except Exception as exc:
                        cell = ws.cell(row=current_row, column=1, value=f"[Image error: {exc}]")
                        cell.font = Font(color="EF4444", italic=True, size=10)
                        current_row += 2

        # Auto-fit column widths
        for col in ws.columns:
            max_len = 0
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 50)

        # Save to bytes
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _load_image_bytes(self, src: str):
        """Return raw image bytes from a base64 data-URL or http/https URL."""
        if src.startswith("data:"):
            # e.g. data:image/png;base64,AAAA...
            _, encoded = src.split(",", 1)
            return base64.b64decode(encoded)
        elif src.startswith("http://") or src.startswith("https://"):
            req = urllib.request.Request(src, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=10) as resp:
                return resp.read()
        return None